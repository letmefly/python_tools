#-*-coding:utf-8 -*-

import openpyxl
import os

################################## 修改如下代码 ######################################
#理论班 
def process_sheets(sheet_name, sheet):
    if sheet_name == '理论课':
        process_lilunke_sheet(sheet)


def process_lilunke_sheet(sheet):
    ####################### 每行处理 #######################
    for row_num in range(2, sheet.max_row+1):
        process_lilunke_row(sheet, row_num)
    
    ####################### 求和处理 #######################
    # sum_title_title(sheet, "列名一", "列名二", "列名一的求和名", "列名二的求和名")
    sum_title_title(sheet, '任课教师姓名', 'Ai工作当量',    '教师', 'Ai工作当量总和')


def process_lilunke_row(sheet, row_num):
    lession_val = 0
    lession_type = get_cell_value(sheet, row_num, '课程类别')
        
    if lession_type == '一般理论课':
        lession_val = 1
    if lession_type == '双语授课课程':
        lession_val = 1.5
    if lession_type == '就业指导课':
        lession_val = 0.85
    
    class_val = 0
    beizhu = get_cell_value(sheet, row_num, '备注')

    if beizhu == '重复班':
        class_val = 0.9
    if beizhu == '非重复班':
        class_val = 1

    
    Ai_val = 0
    student_number = get_cell_value(sheet,row_num, '学生人数')
    keshishu = get_cell_value(sheet,row_num, '课程总学时')

    if student_number < 32:
        student_number = 32

    Ai_val = ((student_number-32)*0.01+1)*keshishu*class_val*lession_val

    print(row_num, lession_type, beizhu, lession_val, class_val, Ai_val)
    set_cell_value(sheet, row_num, 'Ci重复班系数', lession_val)
    set_cell_value(sheet, row_num, 'Ki课程类别系数', class_val)
    set_cell_value(sheet, row_num, 'Ni学生数', student_number)
    set_cell_value(sheet, row_num, 'Ji学时数', keshishu)
    set_cell_value(sheet, row_num, 'Ai工作当量', Ai_val)

    
################################## 修改以上代码 ######################################























def read_excel(excel_full_name):
    wb = openpyxl.load_workbook(excel_full_name)
    first_sheet_name = wb.sheetnames[0]
    sheet = wb[first_sheet_name]
    print("处理"+excel_full_name, "的子表：", sheet.title)
    for row in sheet.rows:
        line = []
        for cell in row:
            line.append(cell.value)
        #print(line)
    return first_sheet_name, sheet

def write_excel(excel_full_name, sheet_name, old_sheet):
    wb = openpyxl.Workbook()
    new_sheet = wb.create_sheet(index=0, title=sheet_name)
    for row_num in range(1, old_sheet.max_row+1):
        for col_num in range(1, old_sheet.max_column+1):
            value = old_sheet.cell(row = row_num, column = col_num).value
            new_sheet.cell(row=row_num, column=col_num, value=value)
    wb.save(excel_full_name)

def is_number(n):
  is_number = True
  try:
    num = float(n)
    is_number = num == num
  except ValueError:
    is_number = False

def get_cell(sheet, row_idx, title_name):
    title_idx = 0
    for col_idx in range(1, sheet.max_column+1):
        cell = sheet.cell(row=1, column=col_idx)
        if cell.value == title_name:
            title_idx = col_idx
            break
    if title_idx == 0:
        print("there is no title:"+title_name)
        exit()
    return sheet.cell(row=row_idx, column=title_idx)

def get_cell_value(sheet, row_idx, title_name):
    cell_value = get_cell(sheet, row_idx, title_name).value
    if cell_value is None or cell_value =="":
        print(row_idx, title_name, "不能为空")
        exit()
    if is_number(cell_value):
        return float(cell_value)
    else:
        return cell_value

def set_cell_value(sheet, row_idx, title_name, value):
    get_cell(sheet, row_idx, title_name).value = value


def sum_title_title(sheet, title1, title2, sum_title1, sum_title2):
    sum_dict = {}
    for row_num in range(2, sheet.max_row+1):
        title1_val = get_cell_value(sheet, row_num, title1)
        title2_val = get_cell_value(sheet, row_num, title2)
        val = sum_dict.get(title1_val, None)
        if val is None:
            sum_dict[title1_val] = 0
        sum_dict[title1_val] = sum_dict[title1_val] + title2_val
    
    row_num = 2
    for key in sum_dict:
        val = sum_dict[key]
        set_cell_value(sheet, row_num, sum_title1, key)
        set_cell_value(sheet, row_num, sum_title2, val)
        row_num = row_num + 1


if __name__ == '__main__':
    filenames = os.listdir("input")
    for filename in filenames:
        excel_path = 'input/' + filename
        file_name = os.path.basename(excel_path)
        sheet_name, sheet = read_excel(excel_path)
        process_sheets(sheet_name, sheet)
        output_excel_name = "output/"+"核算成功_"+file_name
        write_excel(output_excel_name, sheet_name, sheet)
        print(os.path.basename(output_excel_name), '写入成功，在output目录查看')
    
    input("核算成功!")
